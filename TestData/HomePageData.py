import openpyxl


class HomePageData:
    # test_homePage_data = [{"firstName": "ira", "email": "i@i.i", "gender": "Female"},
    #                       {"firstName": "katya", "email": "k@k.k", "gender": "Female"},
    #                       {"firstName": "andrey", "email": "a@a.a", "gender": "Male"}]
    @staticmethod
    def get_test_data(test_case_name):
        book = openpyxl.load_workbook(
            "C:\\Users\\Vernik\\PycharmProjects\\PythomSeleniumFramework\\TestData\\Test_Data.xlsx")
        sheet = book.active
        diction = {}
        for row_element in range(1, sheet.max_row + 1):
            if sheet.cell(row=row_element, column=1).value == test_case_name:
                for colomn_element in range(2, sheet.max_column + 1):
                    diction[sheet.cell(row=1, column=colomn_element).value] = sheet.cell(row=row_element,
                                                                                           column=colomn_element).value
        return [diction]
