import openpyxl

# work with exel file
book = openpyxl.load_workbook("C:\\Users\\Vernik\\PycharmProjects\\PythomSeleniumFramework\\TestData\\Test_Data.xlsx")
# work with exel sheet
sheet1 = book.active
# work with exel cell
cell = sheet1.cell(row=1, column=2)
print(cell.value)
# important to save exel file before run the test

# ________________________________________________________________

# write
sheet1.cell(row=2, column=2).value = 'ira'
# read
print(sheet1.cell(row=2, column=2).value)
print(sheet1['A3'].value)


print(sheet1.max_row)
print(sheet1.max_column)

# ________________________________________________________________

for row_element in range(1,sheet1.max_row+1):
    for colomn_element in range(2, sheet1.max_column+1):
        print(sheet1.cell(row=row_element, column=colomn_element).value, end="\t")
    print("")


# ________________________________________________________________

print("\n ________________________________________________________________")
diction = {}
for row_element in range(1,sheet1.max_row+1):
    if sheet1.cell(row=row_element, column=1).value == "Testcase2":
        for colomn_element in range(2, sheet1.max_column+1):
            diction[sheet1.cell(row=1, column=colomn_element).value] = sheet1.cell(row=row_element, column=colomn_element).value

print(diction)
diction.clear()

# ________________________________________________________________

print("\n ________________________________________________________________")
lictic = []
diction = {}
for row_element in range(1,sheet1.max_row+1):
    for colomn_element in range(2, sheet1.max_column+1):
        diction[sheet1.cell(row=1, column=colomn_element).value] = sheet1.cell(row=row_element, column=colomn_element).value
    lictic.append(diction)
print(diction)
print(lictic)
diction.clear()